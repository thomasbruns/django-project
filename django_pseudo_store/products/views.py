from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import user_passes_test
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q
from .models import Product
from .forms import ProductForm
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from io import BytesIO
import csv

def product_list(request):
    # Sorting
    sort_by = request.GET.get('sort', 'created_at')  # Default sorting by 'created_at'

    current_sort = request.GET.get('current_sort', None)
    if current_sort == sort_by:
        # If it is the same, reverse the order
        sort_by = '-' + sort_by
        current_sort = '-' + current_sort
    else:
        # If it is not the same, set the current_sort parameter to the sort_by value
        current_sort = sort_by
    
    products = Product.objects.filter(is_deleted=False).order_by(sort_by)

    # Filtering
    search_query = request.GET.get('search', '')
    if search_query:
        products = products.filter(Q(title__icontains=search_query))

    # Pagination
    paginator = Paginator(products, 10)  # Display 10 products per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'products': page_obj,
        'sort_by': sort_by,
        'search_query': search_query,
        'current_sort': current_sort,
    }
    
    return render(request, 'products/product_list.html', context)


@user_passes_test(lambda u: u.has_perm('products.view_product'))
def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk, is_deleted=False)  # Exclude deleted products
    context = {'product': product}
    return render(request, 'products/product_detail.html', context)


@user_passes_test(lambda u: u.has_perm('products.add_product'))
def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('products:product_list')
    else:
        form = ProductForm()
    context = {'form': form}
    return render(request, 'products/product_create.html', context)


@user_passes_test(lambda u: u.has_perm('products.change_product'))
def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk, is_deleted=False)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            return redirect('products:product_list')
    else:
        form = ProductForm(instance=product)
    context = {'form': form}
    return render(request, 'products/product_update.html', context)


@user_passes_test(lambda u: u.has_perm('products.delete_product'))
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.is_deleted = True  # Soft delete the product
        product.save()
        return redirect('products:product_list')
    context = {'product': product}
    return render(request, 'products/product_delete.html', context)


def export_products(request):
    format = request.GET.get('format', 'csv')  # Get the export format from the query parameters

    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="products.csv"'

        writer = csv.writer(response)
        writer.writerow(['ID', 'Title', 'Description', 'Price', 'Created At', 'Modified At'])

        products = Product.objects.filter(is_deleted=False).order_by('created_at')
        for product in products:
            writer.writerow([product.id, product.title, product.description, product.price, product.created_at, product.modified_at])

    elif format == 'excel':
        # Export as Excel
        # Create a new Workbook
        wb = Workbook()
        ws = wb.active

        # Set the headers
        headers = ['ID', 'Title', 'Description', 'Price', 'Created At', 'Modified At']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws['{}1'.format(col_letter)] = header

        # Fetch the products
        products = Product.objects.filter(is_deleted=False).order_by('created_at')

        # Populate the rows with product data
        for row_num, product in enumerate(products, 2):
            created_at = product.created_at.astimezone(timezone.utc).replace(tzinfo=None)
            modified_at = product.modified_at.astimezone(timezone.utc).replace(tzinfo=None)

            ws.cell(row=row_num, column=1, value=product.id)
            ws.cell(row=row_num, column=2, value=product.title)
            ws.cell(row=row_num, column=3, value=product.description)
            ws.cell(row=row_num, column=4, value=product.price)
            ws.cell(row=row_num, column=5, value=created_at)
            ws.cell(row=row_num, column=6, value=modified_at)

        # Create an in-memory file-like object for storing the workbook
        file_stream = BytesIO()

        # Save the workbook to the file-like object
        wb.save(file_stream)

        # Set the appropriate response headers for Excel file download
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="products.xlsx"'

        # Move the file pointer to the beginning of the file-like object
        file_stream.seek(0)

        # Copy the workbook file-like object to the response
        response.write(file_stream.read())

    else:
        # Invalid format, return a 400 Bad Request response
        response = HttpResponseBadRequest("Invalid export format.")

    return response

